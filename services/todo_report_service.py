import io
from typing import List
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from models.todo_task import ToDoTask


class TodoReportService:#передать на io files
    @staticmethod
    def create_excel_buffer(todos: List[ToDoTask]) -> io.BytesIO:
        """
        Создает Excel файл с todo задачами пользователя
        """
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Todo Tasks"
        
        # Заголовки
        headers = ["ID", "Название", "Описание", "Дата создания", "Дата изменения"]
        
        # Стилизация заголовков
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Добавляем заголовки
        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Добавляем данные
        for row_num, todo in enumerate(todos, 2):
            worksheet.cell(row=row_num, column=1, value=str(todo.id))
            worksheet.cell(row=row_num, column=2, value=todo.title)
            worksheet.cell(row=row_num, column=3, value=todo.description or "")
            worksheet.cell(row=row_num, column=4, value=todo.created_at.strftime("%d.%m.%Y %H:%M") if todo.created_at else "")
            worksheet.cell(row=row_num, column=5, value=todo.updated_at.strftime("%d.%m.%Y %H:%M") if todo.updated_at else "")
        
        # Сохраняем в BytesIO
        excel_buffer = io.BytesIO()
        workbook.save(excel_buffer)
        excel_buffer.seek(0)
        
        return excel_buffer