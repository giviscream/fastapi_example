import io
from typing import AsyncGenerator, List
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from models.todo_task import ToDoTask
from resources.constants import (
    BODY_ROW_NUM,
    COLUMNS_WIDTH,
    HEADER_ALIGNMENT,
    HEADER_FILL_KWARGS,
    HEADER_FONT,
    HEADERS_COL_NUM,
    HEADERS_ROW_NUM,
    REPORT_CHUNK_SIZE,
    REPORT_DATETIME_FORMAT,
    REPORT_HEADERS,
    REPORT_TITLE,
)


class TodoReportService:
    def _create_excel_report_buffer(self, todos: List[ToDoTask]) -> io.BytesIO:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = REPORT_TITLE

        header_font = Font(**HEADER_FONT)
        header_fill = PatternFill(**HEADER_FILL_KWARGS)
        header_alignment = Alignment(**HEADER_ALIGNMENT)

        for col_num, header in enumerate(
            iterable=REPORT_HEADERS, start=HEADERS_COL_NUM
        ):
            cell = worksheet.cell(row=HEADERS_ROW_NUM, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            
            col_letter = get_column_letter(col_num)
            worksheet.column_dimensions[col_letter].width = COLUMNS_WIDTH

        for row_num, todo in enumerate(iterable=todos, start=BODY_ROW_NUM):
            worksheet.cell(
                row=row_num,
                column=REPORT_HEADERS.index("ID") + 1,
                value=str(object=todo.id),
            )
            worksheet.cell(
                row=row_num, column=REPORT_HEADERS.index("Название") + 1, value=todo.title
            )
            worksheet.cell(
                row=row_num,
                column=REPORT_HEADERS.index("Описание") + 1,
                value=todo.description or "",
            )
            worksheet.cell(
                row=row_num,
                column=REPORT_HEADERS.index("Дата создания") + 1,
                value=(
                    todo.created_at.strftime(REPORT_DATETIME_FORMAT)
                    if todo.created_at
                    else ""
                ),
            )
            worksheet.cell(
                row=row_num,
                column=REPORT_HEADERS.index("Дата изменения") + 1,
                value=(
                    todo.updated_at.strftime(REPORT_DATETIME_FORMAT)
                    if todo.updated_at
                    else ""
                ),
            )

        buf = io.BytesIO()
        workbook.save(filename=buf)
        buf.seek(0)
        return buf

    async def get_report_chunks(
        self,
        todos: List[ToDoTask],
        chunk_size: int = REPORT_CHUNK_SIZE,
    ) -> AsyncGenerator[bytes, None]:
        buf: io.BytesIO = self._create_excel_report_buffer(todos=todos)

        while chunk := buf.read(chunk_size):
            yield chunk

        buf.close()
